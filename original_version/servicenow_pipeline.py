#!/usr/bin/env python3
"""Lightweight ServiceNow HR case processing pipeline.

Inputs:
- Excel or CSV exports from ServiceNow

Outputs:
- enriched_cases.csv
- pipeline_summary.json
- pipeline_summary.xlsx

Designed for first-prototype analytics and Streamlit ingestion.
"""
from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

STATE_MAP = {
    "work in progress": "Work in Progress",
    "awaiting acceptance": "Awaiting Acceptance",
    "closed complete": "Closed Complete",
    "closed incomplete": "Closed Incomplete",
    "new": "New",
    "resolved": "Resolved",
}

PRIORITY_ORDER = {
    "1 - critical": 1,
    "2 - high": 2,
    "3 - moderate": 3,
    "4 - low": 4,
    "5 - planning": 5,
}

CATEGORY_RULES: list[tuple[str, list[str]]] = [
    ("Immigration", ["visa", "permit", "residence", "entry", "cancellation", "uae"]),
    (
        "Relocation & Shipment",
        ["shipment", "shipping", "relocation", "repatriation", "movers", "safe journey", "abu dhabi delight", "academic shipment"],
    ),
    ("Training & Learning", ["training", "course", "ilearn", "coaching", "professional development", "learning"]),
    ("Employment & Contracts", ["contract", "extension", "ending", "fixed term", "employment", "temporary work"]),
    ("Documents & Certificates", ["certificate", "document", "legalization", "upload", "copy"]),
    ("Travel & Expenses", ["travel", "expense", "ticketing", "booking"]),
    ("General Inquiry & Follow-up", ["inquiry", "follow up", "follow-up", "update", "status check", "re:", "fwd:"]),
]

URGENCY_TERMS = ["urgent", "asap", "immediately", "today", "soonest", "priority"]
FRUSTRATION_TERMS = ["follow up", "follow-up", "still waiting", "no response", "delayed", "delay", "pending too long"]
BLOCKED_TERMS = ["awaiting", "pending", "approval", "clearance", "acceptance", "blocked"]
KB_TERMS = ["how to", "procedure", "policy", "guideline", "where can i", "what is the process"]


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", text.strip().lower()).strip("_")


def read_servicenow_file(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(path)
    if suffix == ".csv":
        for encoding in ("utf-8", "utf-8-sig", "latin1", "cp1252"):
            try:
                return pd.read_csv(path, encoding=encoding)
            except UnicodeDecodeError:
                continue
        raise ValueError(f"Unable to read CSV with supported encodings: {path}")
    raise ValueError(f"Unsupported file type: {path.suffix}")


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [_slug(c) for c in out.columns]
    return out


def clean_text(text: object) -> str:
    if pd.isna(text):
        return ""
    value = str(text).strip()
    value = re.sub(r"^(re|fw|fwd):\s*", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_state(text: object) -> str | None:
    if pd.isna(text):
        return None
    value = str(text).strip()
    key = value.lower()
    return STATE_MAP.get(key, value.title())


def normalize_priority(text: object) -> str | None:
    if pd.isna(text):
        return None
    value = str(text).strip()
    if re.match(r"^[1-5]\s*-\s*", value):
        parts = value.split("-", 1)
        return f"{parts[0].strip()} - {parts[1].strip().title()}"
    return value.title()


def categorize_case(text: str) -> str:
    text_l = text.lower()
    for category, keywords in CATEGORY_RULES:
        if any(k in text_l for k in keywords):
            return category
    return "Other"


def classify_operational_signal(text: str) -> str:
    text_l = text.lower()
    if any(k in text_l for k in FRUSTRATION_TERMS):
        return "Frustrated / Follow-up"
    if any(k in text_l for k in URGENCY_TERMS):
        return "Urgent"
    if any(k in text_l for k in BLOCKED_TERMS):
        return "Blocked / Waiting"
    if any(k in text_l for k in KB_TERMS):
        return "KB Opportunity"
    return "Normal"


def recommend_action(category: str, signal: str, state: str | None) -> str:
    if signal == "KB Opportunity":
        return "Create or update KB article"
    if signal in {"Frustrated / Follow-up", "Blocked / Waiting"}:
        return "Review stuck workflow and next-owner handoff"
    if category in {"Immigration", "Relocation & Shipment", "Documents & Certificates"}:
        return "Review process steps and checklist completeness"
    if category == "Training & Learning":
        return "Standardize common request handling"
    if state == "Awaiting Acceptance":
        return "Follow up with requester or approver"
    return "No immediate action"


def first_present(df: pd.DataFrame, candidates: Iterable[str]) -> str | None:
    for name in candidates:
        if name in df.columns:
            return name
    return None


@dataclass
class PipelineResult:
    enriched: pd.DataFrame
    summary: dict


def process_dataframe(df_raw: pd.DataFrame, reference_time: str | None = None) -> PipelineResult:
    df = normalize_columns(df_raw)

    number_col = first_present(df, ["number", "case_number", "ticket_number"])
    desc_col = first_present(df, ["short_description", "description", "full_description"])
    state_col = first_present(df, ["state", "status"])
    priority_col = first_present(df, ["priority"])
    created_col = first_present(df, ["created", "opened_at", "opened", "sys_created_on"])

    if not number_col or not desc_col:
        raise ValueError("Dataset must contain at least a case number column and a description column.")

    enriched = df.copy()
    enriched["case_number"] = enriched[number_col].astype(str).str.strip()
    enriched["short_description_clean"] = enriched[desc_col].map(clean_text)

    if state_col:
        enriched["state_normalized"] = enriched[state_col].map(normalize_state)
    else:
        enriched["state_normalized"] = None

    if priority_col:
        enriched["priority_normalized"] = enriched[priority_col].map(normalize_priority)
        enriched["priority_rank"] = (
            enriched["priority_normalized"].str.lower().map(PRIORITY_ORDER).astype("Int64")
        )
    else:
        enriched["priority_normalized"] = None
        enriched["priority_rank"] = pd.Series([pd.NA] * len(enriched), dtype="Int64")

    if created_col:
        enriched["created_ts"] = pd.to_datetime(enriched[created_col], errors="coerce")
        ref = pd.Timestamp.now(tz=None) if reference_time is None else pd.Timestamp(reference_time)
        enriched["case_age_days"] = (ref - enriched["created_ts"]).dt.total_seconds() / 86400.0
        enriched["case_age_days"] = enriched["case_age_days"].round(1)
        enriched["created_date"] = enriched["created_ts"].dt.date
        enriched["created_month"] = enriched["created_ts"].dt.to_period("M").astype(str)
    else:
        enriched["created_ts"] = pd.NaT
        enriched["case_age_days"] = pd.NA
        enriched["created_date"] = pd.NA
        enriched["created_month"] = pd.NA

    enriched["request_category"] = enriched["short_description_clean"].map(categorize_case)
    enriched["operational_signal"] = enriched["short_description_clean"].map(classify_operational_signal)
    enriched["recommended_action"] = enriched.apply(
        lambda row: recommend_action(
            row.get("request_category", "Other"),
            row.get("operational_signal", "Normal"),
            row.get("state_normalized"),
        ),
        axis=1,
    )

    enriched["is_open_case"] = ~enriched["state_normalized"].fillna("").str.lower().isin(
        {"closed complete", "closed incomplete", "resolved"}
    )
    enriched["is_aging_7d"] = enriched["case_age_days"].fillna(-1).ge(7)
    enriched["is_aging_30d"] = enriched["case_age_days"].fillna(-1).ge(30)

    before_dedup = len(enriched)
    enriched = enriched.drop_duplicates(subset=["case_number"], keep="first").reset_index(drop=True)
    deduped = before_dedup - len(enriched)

    summary = {
        "row_count": int(len(enriched)),
        "duplicate_rows_removed": int(deduped),
        "columns_available": list(df.columns),
        "date_range": {
            "min_created": None if enriched["created_ts"].isna().all() else str(enriched["created_ts"].min()),
            "max_created": None if enriched["created_ts"].isna().all() else str(enriched["created_ts"].max()),
        },
        "state_counts": enriched["state_normalized"].fillna("Unknown").value_counts().to_dict(),
        "priority_counts": enriched["priority_normalized"].fillna("Unknown").value_counts().to_dict(),
        "category_counts": enriched["request_category"].value_counts().to_dict(),
        "signal_counts": enriched["operational_signal"].value_counts().to_dict(),
        "aging": {
            "older_than_7_days": int(enriched["is_aging_7d"].sum()),
            "older_than_30_days": int(enriched["is_aging_30d"].sum()),
            "average_age_days": None if enriched["case_age_days"].isna().all() else round(float(enriched["case_age_days"].mean()), 1),
        },
    }
    return PipelineResult(enriched=enriched, summary=summary)


def write_summary_workbook(result: PipelineResult, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(["Metric", "Value"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    rows = [
        ("Rows after dedupe", result.summary["row_count"]),
        ("Duplicates removed", result.summary["duplicate_rows_removed"]),
        ("Avg age (days)", result.summary["aging"]["average_age_days"]),
        ("Older than 7 days", result.summary["aging"]["older_than_7_days"]),
        ("Older than 30 days", result.summary["aging"]["older_than_30_days"]),
        ("Min created", result.summary["date_range"]["min_created"]),
        ("Max created", result.summary["date_range"]["max_created"]),
    ]
    for row in rows:
        ws.append(list(row))

    def add_section(title: str, mapping: dict, start_row: int) -> int:
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(bold=True)
        ws.cell(row=start_row + 1, column=1, value="Label")
        ws.cell(row=start_row + 1, column=2, value="Count")
        for cell in (ws.cell(row=start_row + 1, column=1), ws.cell(row=start_row + 1, column=2)):
            cell.fill = header_fill
            cell.font = header_font
        r = start_row + 2
        for k, v in mapping.items():
            ws.cell(row=r, column=1, value=k)
            ws.cell(row=r, column=2, value=v)
            r += 1
        return r + 1

    next_row = 10
    next_row = add_section("State counts", result.summary["state_counts"], next_row)
    next_row = add_section("Priority counts", result.summary["priority_counts"], next_row)
    next_row = add_section("Category counts", result.summary["category_counts"], next_row)
    add_section("Operational signals", result.summary["signal_counts"], next_row)

    for idx, width in {1: 28, 2: 24}.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws2 = wb.create_sheet("enriched_preview")
    preview_cols = [
        "case_number",
        "short_description_clean",
        "state_normalized",
        "priority_normalized",
        "request_category",
        "operational_signal",
        "recommended_action",
        "case_age_days",
    ]
    preview = result.enriched[[c for c in preview_cols if c in result.enriched.columns]].head(25)
    ws2.append(list(preview.columns))
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in preview.itertuples(index=False):
        ws2.append(list(row))
    for idx, col in enumerate(preview.columns, start=1):
        max_len = max(len(str(col)), *(len(str(v)) for v in preview[col].fillna("")))
        ws2.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 14), 40)

    wb.save(output_path)


def run_pipeline(input_path: Path, output_dir: Path, reference_time: str | None = None) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    raw = read_servicenow_file(input_path)
    result = process_dataframe(raw, reference_time=reference_time)

    enriched_csv = output_dir / "enriched_cases.csv"
    summary_json = output_dir / "pipeline_summary.json"
    summary_xlsx = output_dir / "pipeline_summary.xlsx"

    result.enriched.to_csv(enriched_csv, index=False)
    summary_json.write_text(json.dumps(result.summary, indent=2), encoding="utf-8")
    write_summary_workbook(result, summary_xlsx)

    return {
        "enriched_csv": enriched_csv,
        "summary_json": summary_json,
        "summary_xlsx": summary_xlsx,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Process ServiceNow case export into an enriched analytics dataset.")
    parser.add_argument("input_file", type=Path, help="Path to ServiceNow CSV/XLSX export")
    parser.add_argument("--output-dir", type=Path, default=Path("./output"), help="Directory for generated files")
    parser.add_argument("--reference-time", type=str, default=None, help="Optional reference timestamp for age calculations")
    args = parser.parse_args()

    outputs = run_pipeline(args.input_file, args.output_dir, reference_time=args.reference_time)
    print("Generated files:")
    for name, path in outputs.items():
        print(f"- {name}: {path}")


if __name__ == "__main__":
    main()
